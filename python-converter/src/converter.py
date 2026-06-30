from openpyxl import load_workbook

from src.models import Emenda, Status
from src.formatter import (
    format_emenda,
    format_name,
    format_date,
    format_money,
    parse_money
)

from src.validator import (
    validate_row
)

from src.documents import process_documents

from utils.logger import get_logger

logger = get_logger()

def convert_file(path):
    workbook = load_workbook(path)
    sheet = workbook.active

    headers = [
        cell.value
        for cell in sheet[1]
    ]

    result = []


    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        data = dict(zip(headers, row))

        errors = validate_row(data)

        if errors:
            logger.warning(
                f'Linha {row_number + 2} no arquivo "{path.name}" ignorada. Motivo(s): {errors}'
            )
            continue
        
        numero_emenda = format_emenda(data["Número"])
        documentos = process_documents(numero_emenda)

        emenda = Emenda(
            numeroEmenda=format_emenda(data["Número"]),
            
            parlamentarAutor=format_name(data["Parlamentar Autor"]),
            
            secretaria=data["Secretaria"],

            beneficiario=data["Beneficiário"] or "",
            objeto=data["Objeto"] or "",
            
            tipo=data["Tipo"],

            valorPrevisto=format_money(
                data["Valor Previsto"]
            ),
            valorPrevistoNum=parse_money(
                data["Valor Previsto"]
            ),
            
            valorEmpenhado=format_money(
                data["Valor Empenhado"]
            ),
            valorEmpenhadoNum=parse_money(
                data["Valor Empenhado"]
            ),
            
            valorLiquidado=format_money(
                data["Valor Liquidado"]
            ),
            valorLiquidadoNum=parse_money(
                data["Valor Liquidado"]
            ),

            valorPago=format_money(
                data["Valor Pago"]
            ),
            valorPagoNum=parse_money(
                data["Valor Pago"]
            ),

            status=Status(data["Status"]),

            dataEstimadaConclusao=format_date(
                data["Data Estimada de Conclusão"]
            ),

            planoTrabalho=bool(
                data["Plano de Trabalho"]
            ),

            notaFiscal=bool(
                data["Nota Fiscal"]
            ),
            
            hasDocuments=bool(documentos),
            
            documents=documentos
        )

        result.append(emenda)

    return result